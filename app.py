from __future__ import annotations

import argparse
import json
import mimetypes
import os
import shutil
import tempfile
import warnings
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import unquote, urlparse

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from channel_report_generator import (
    AXIO_FILE as DEFAULT_AXIO_FILE,
    OUTPUT_FILE as CHANNEL_REPORT_FILE,
    RETAIL_FILE as DEFAULT_RETAIL_FILE,
    generate_channel_report,
)

import sys
from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

with warnings.catch_warnings():
    warnings.simplefilter("ignore", DeprecationWarning)
    import cgi

STATIC_DIR = BASE_DIR / "static"
OUTPUT_DIR = Path("/tmp") if os.getenv("VERCEL") else BASE_DIR

DEFAULT_SERVICE_FILE = BASE_DIR / "mi_smart_report (6).csv"
DEFAULT_SERVICE_MASTER_FILE = BASE_DIR / "current_service_master.xlsx"
DEFAULT_CHANNEL_MASTER_FILE = BASE_DIR / "Master April'26.xlsb"
FINAL_REPORT_FILE = OUTPUT_DIR / "final_report.xlsx"
ZONAL_REPORT_FILE = OUTPUT_DIR / "zonal_report.xlsx"

APP_NAME = "Xiaomi Daily Report Engine"
EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

REQUIRED_SERVICE_COLUMNS = {
    "PAYMENT STATUS",
    "ASC Code",
    "CUSTOMER PRICE",
}

NEW_SERVICE_MASTER_COLUMNS = {"Agency_Code", "Agency_Name", "Region"}
LEGACY_SERVICE_MASTER_COLUMNS = {"ASC_Code", "ASC_Name_BI", "Zone", "State"}


def normalise_truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().upper() in {"TRUE", "1", "YES", "Y", "PAID"}


def clean_dimension(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("Blank").astype(str).str.strip()
    return cleaned.mask(cleaned.eq("") | cleaned.str.lower().eq("nan"), "Blank")


def read_master_workbook(path: Path, sheet_name: int | str = 0) -> pd.DataFrame:
    suffix = path.suffix.lower()
    engine = "pyxlsb" if suffix == ".xlsb" else None
    return pd.read_excel(path, sheet_name=sheet_name, engine=engine)


def validate_columns(frame: pd.DataFrame, required: set[str], label: str) -> None:
    missing = sorted(required.difference(frame.columns))
    if missing:
        missing_text = ", ".join(missing)
        raise ValueError(f"{label} is missing required column(s): {missing_text}")


def ordered_with_blank_last(values: pd.Series) -> list[str]:
    unique_values = list(dict.fromkeys(values.tolist()))
    ordered = [value for value in unique_values if value != "Blank"]
    if "Blank" in unique_values:
        ordered.append("Blank")
    return ordered


def normalise_numeric_cell(value: object) -> int:
    if value in (None, "") or pd.isna(value):
        return 0
    return int(round(float(value)))


def fill_empty_numeric_cells(
    worksheet: object,
    *,
    numeric_columns: tuple[int, ...],
    min_row: int,
) -> None:
    for row in worksheet.iter_rows(min_row=min_row):
        for column in numeric_columns:
            cell = row[column - 1]
            cell.value = normalise_numeric_cell(cell.value)


def normalise_service_master(master: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    if NEW_SERVICE_MASTER_COLUMNS.issubset(master.columns):
        state_column = "State-2" if "State-2" in master.columns else "State"
        if state_column not in master.columns:
            raise ValueError("Service master is missing required column: State or State-2")
        normalised = pd.DataFrame(
            {
                "ServiceCode": master["Agency_Code"],
                "Region": master["Region"],
                "State": master[state_column],
                "ServiceCenter": master["Agency_Name"],
            }
        )
        return normalised, "Agency_Code"

    if LEGACY_SERVICE_MASTER_COLUMNS.issubset(master.columns):
        normalised = pd.DataFrame(
            {
                "ServiceCode": master["ASC_Code"],
                "Region": master["Zone"],
                "State": master["State"],
                "ServiceCenter": master["ASC_Name_BI"],
            }
        )
        return normalised, "ASC_Code"

    required = " / ".join(
        [
            ", ".join(sorted(NEW_SERVICE_MASTER_COLUMNS)),
            ", ".join(sorted(LEGACY_SERVICE_MASTER_COLUMNS)),
        ]
    )
    raise ValueError(f"Master workbook is missing service master columns. Expected one of: {required}")


def build_final_rows(report_df: pd.DataFrame) -> list[dict[str, object]]:
    final_rows: list[dict[str, object]] = []
    grand_unit_total = 0
    grand_gwp_total = 0.0

    for region in ordered_with_blank_last(report_df["Region"]):
        region_df = report_df[report_df["Region"] == region]
        region_unit_total = 0
        region_gwp_total = 0.0
        first_region_row = True

        for state in ordered_with_blank_last(region_df["State"]):
            state_df = region_df[region_df["State"] == state]
            state_unit_total = int(state_df["Unit"].sum())
            state_gwp_total = float(state_df["GWP"].sum())

            region_unit_total += state_unit_total
            region_gwp_total += state_gwp_total
            grand_unit_total += state_unit_total
            grand_gwp_total += state_gwp_total
            first_state_row = True

            for _, row in state_df.iterrows():
                final_rows.append(
                    {
                        "Region": region if first_region_row else "",
                        "State": state if first_state_row else "",
                        "Service Center Name": row["ServiceCenter"],
                        "Unit": int(row["Unit"]),
                        "GWP": int(round(float(row["GWP"]))),
                    }
                )
                first_region_row = False
                first_state_row = False

            final_rows.append(
                {
                    "Region": "",
                    "State": f"{state} Total",
                    "Service Center Name": "",
                    "Unit": state_unit_total,
                    "GWP": int(round(state_gwp_total)),
                }
            )

        final_rows.append(
            {
                "Region": f"{region} Total",
                "State": "",
                "Service Center Name": "",
                "Unit": int(region_unit_total),
                "GWP": int(round(region_gwp_total)),
            }
        )

    final_rows.append(
        {
            "Region": "Grand Total",
            "State": "",
            "Service Center Name": "",
            "Unit": int(grand_unit_total),
            "GWP": int(round(grand_gwp_total)),
        }
    )
    return final_rows


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    worksheet = workbook.active

    worksheet.insert_rows(1)
    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = "Xiaomi DSR 1st May To Till Date"

    title_fill = PatternFill("solid", fgColor="ED6B2A")
    header_fill = PatternFill("solid", fgColor="A9CBE8")
    title_font = Font(color="000000", bold=True)
    header_font = Font(color="000000", bold=True)
    total_fill = PatternFill("solid", fgColor="FFF3E8")
    grand_fill = PatternFill("solid", fgColor="F15A24")
    border_color = "808080"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 23

    for cell in worksheet[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    fill_empty_numeric_cells(worksheet, numeric_columns=(4, 5), min_row=3)

    for row in worksheet.iter_rows(min_row=3):
        region = str(row[0].value or "")
        state = str(row[1].value or "")
        is_grand_total = region == "Grand Total"
        is_total = is_grand_total or region.endswith("Total") or state.endswith("Total")

        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if cell.column in {4, 5}:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if is_total:
                cell.fill = grand_fill if is_grand_total else total_fill
                cell.font = Font(color="FFFFFF" if is_grand_total else "111827", bold=True)

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:E{worksheet.max_row}"

    widths = {
        "A": 18,
        "B": 24,
        "C": 42,
        "D": 14,
        "E": 16,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    workbook.save(path)


def style_channel_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    worksheet = workbook.active

    worksheet.insert_rows(1, amount=2)
    worksheet.merge_cells("A1:I1")
    worksheet["A1"] = "Xiaomi DSR 1st May To Till Date"
    worksheet.merge_cells("D2:E2")
    worksheet.merge_cells("F2:G2")
    worksheet["D2"] = "AXIO"
    worksheet["F2"] = "Retail"
    worksheet["H2"] = "Total Unit"
    worksheet["I2"] = "Total GWP"

    header_labels = {
        "A3": "State",
        "B3": "DistributorName",
        "C3": "RetailerName",
        "D3": "Unit",
        "E3": "GWP",
        "F3": "Unit",
        "G3": "GWP",
        "H3": "Total Unit",
        "I3": "Total GWP",
    }
    for cell_ref, value in header_labels.items():
        worksheet[cell_ref] = value

    title_fill = PatternFill("solid", fgColor="ED6B2A")
    header_fill = PatternFill("solid", fgColor="A9CBE8")
    title_font = Font(color="000000", bold=True)
    header_font = Font(color="000000", bold=True)
    total_fill = PatternFill("solid", fgColor="FFF3E8")
    grand_fill = PatternFill("solid", fgColor="F15A24")
    border_color = "808080"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 23

    for cell in worksheet[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    fill_empty_numeric_cells(worksheet, numeric_columns=(4, 5, 6, 7, 8, 9), min_row=4)

    for row in worksheet.iter_rows(min_row=4):
        state = str(row[0].value or "")
        distributor = str(row[1].value or "")
        is_grand_total = state == "Grand Total"
        is_total = is_grand_total or state.endswith("Total") or distributor.endswith("Total")

        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if cell.column in {4, 5, 6, 7, 8, 9}:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if is_total:
                cell.fill = grand_fill if is_grand_total else total_fill
                cell.font = Font(color="FFFFFF" if is_grand_total else "111827", bold=True)

    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = f"A3:I{worksheet.max_row}"

    widths = {
        "A": 22,
        "B": 34,
        "C": 52,
        "D": 12,
        "E": 14,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 14,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    workbook.save(path)


def generate_service_report(service_path: Path, master_path: Path) -> dict[str, object]:
    service = pd.read_csv(service_path)
    validate_columns(service, REQUIRED_SERVICE_COLUMNS, "Daily service report")

    master = read_master_workbook(master_path)
    try:
        master, code_column = normalise_service_master(master)
    except ValueError:
        if master_path.suffix.lower() != ".xlsb":
            raise
        master = read_master_workbook(master_path, sheet_name=1)
        master, code_column = normalise_service_master(master)

    service_input_rows = int(len(service))
    service = service.copy()
    service["ASC Code"] = service["ASC Code"].astype(str).str.strip()
    service["CUSTOMER PRICE"] = pd.to_numeric(service["CUSTOMER PRICE"], errors="coerce").fillna(0)
    service = service[service["PAYMENT STATUS"].map(normalise_truthy)].copy()

    master = master.copy()
    master["ServiceCode"] = master["ServiceCode"].astype(str).str.strip()

    merged = service.merge(
        master,
        left_on="ASC Code",
        right_on="ServiceCode",
        how="left",
    )

    unmatched_rows = int(merged["ServiceCode"].isna().sum())
    merged["Region"] = clean_dimension(merged["Region"])
    merged["State"] = clean_dimension(merged["State"])
    merged["ServiceCenter"] = clean_dimension(merged["ServiceCenter"])

    pivot_df = pd.pivot_table(
        merged,
        index=["Region", "State", "ServiceCenter"],
        values="CUSTOMER PRICE",
        aggfunc=["count", "sum"],
        margins=True,
        margins_name="Total",
    )
    pivot_df.columns = ["Count of PAYMENT STATUS", "Sum of CUSTOMER PRICE"]
    pivot_df.to_excel(ZONAL_REPORT_FILE)

    report_df = (
        merged.groupby(["Region", "State", "ServiceCenter"], as_index=False)
        .agg(Unit=("PAYMENT STATUS", "count"), GWP=("CUSTOMER PRICE", "sum"))
        .sort_values(["Region", "State", "ServiceCenter"], kind="stable")
    )

    final_report = pd.DataFrame(build_final_rows(report_df))
    final_report.to_excel(FINAL_REPORT_FILE, index=False, sheet_name="Daily Report")
    style_workbook(FINAL_REPORT_FILE)

    total_units = int(final_report.loc[final_report["Region"] == "Grand Total", "Unit"].iloc[0])
    total_gwp = int(final_report.loc[final_report["Region"] == "Grand Total", "GWP"].iloc[0])
    service_center_count = int(report_df["ServiceCenter"].nunique())

    generated_at = datetime.now().strftime("%d %b %Y, %I:%M %p")
    return {
        "label": "Service",
        "summary": {
            "input_rows": service_input_rows,
            "paid_rows": int(len(service)),
            "unmatched_rows": unmatched_rows,
            "regions": int(report_df["Region"].nunique()),
            "service_centers": service_center_count,
            "total_units": total_units,
            "total_gwp": total_gwp,
            "generated_at": generated_at,
            "master_key": code_column,
        },
        "preview": final_report.head(40).to_dict(orient="records"),
        "columns": final_report.columns.tolist(),
        "downloads": {
            "final_report": "final_report",
            "zonal_report": "zonal_report",
        },
    }


def generate_channel_payload(axio_path: Path, retail_path: Path, master_path: Path) -> dict[str, object]:
    channel_report = generate_channel_report(
        axio_path=axio_path,
        retail_path=retail_path,
        master_path=master_path,
        output_path=CHANNEL_REPORT_FILE,
    )
    style_channel_workbook(CHANNEL_REPORT_FILE)

    channel_grand = channel_report.loc[channel_report["State"] == "Grand Total"].iloc[0]
    channel_state_total_mask = (
        channel_report["State"].astype(str).str.endswith(" Total")
        & channel_report["State"].ne("Grand Total")
    )
    channel_distributor_total_mask = channel_report["DistributorName"].astype(str).str.endswith(
        " Total"
    )
    channel_detail_mask = (
        channel_report["State"].ne("Grand Total")
        & ~channel_state_total_mask
        & ~channel_distributor_total_mask
    )
    generated_at = datetime.now().strftime("%d %b %Y, %I:%M %p")

    channel_payload = {
        "label": "Retail + Axio",
        "summary": {
            "axio_units": int(channel_grand["AXIO Unit"]),
            "axio_gwp": int(channel_grand["AXIO GWP"]),
            "retail_units": int(channel_grand["Retail Unit"]),
            "retail_gwp": int(channel_grand["Retail GWP"]),
            "states": int(channel_state_total_mask.sum()),
            "distributors": int(channel_distributor_total_mask.sum()),
            "stores": int(channel_detail_mask.sum()),
            "total_units": int(channel_grand["Total Unit"]),
            "total_gwp": int(channel_grand["Total GWP"]),
            "generated_at": generated_at,
        },
        "preview": channel_report.head(40).to_dict(orient="records"),
        "columns": channel_report.columns.tolist(),
        "downloads": {
            "channel_report": "channel_report",
        },
    }
    return channel_payload


def generate_reports(
    report_type: str,
    service_path: Path | None = None,
    service_master_path: Path | None = None,
    axio_path: Path | None = None,
    retail_path: Path | None = None,
    channel_master_path: Path | None = None,
) -> dict[str, object]:
    if report_type == "service":
        if service_path is None or service_master_path is None:
            raise ValueError("Service report requires service file and master workbook.")
        service_payload = generate_service_report(service_path, service_master_path)
        return {
            "active_report": "service",
            "reports": {"service": service_payload},
            "summary": service_payload["summary"],
            "preview": service_payload["preview"],
            "columns": service_payload["columns"],
            "downloads": service_payload["downloads"],
        }

    if report_type == "channel":
        if axio_path is None or retail_path is None or channel_master_path is None:
            raise ValueError("Channel report requires AXIO, retail, and master files.")
        channel_payload = generate_channel_payload(axio_path, retail_path, channel_master_path)
        return {
            "active_report": "channel",
            "reports": {"channel": channel_payload},
            "summary": channel_payload["summary"],
            "preview": channel_payload["preview"],
            "columns": channel_payload["columns"],
            "downloads": channel_payload["downloads"],
        }

    raise ValueError("Unknown report type.")


def file_status(path: Path) -> dict[str, object]:
    if not path.exists():
        return {"exists": False, "name": path.name}
    stat = path.stat()
    return {
        "exists": True,
        "name": path.name,
        "size": stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%d %b %Y, %I:%M %p"),
    }


def write_json(handler: BaseHTTPRequestHandler, payload: dict[str, object], status: int = 200) -> None:
    body = json.dumps(payload, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def serve_file(
    handler: BaseHTTPRequestHandler,
    path: Path,
    content_type: str | None = None,
    attachment_name: str | None = None,
) -> None:
    if not path.exists() or not path.is_file():
        write_json(handler, {"error": "File not found."}, HTTPStatus.NOT_FOUND)
        return

    body = path.read_bytes()
    guessed_type = content_type or mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    handler.send_response(HTTPStatus.OK)
    handler.send_header("Content-Type", guessed_type)
    handler.send_header("Content-Length", str(len(body)))
    if attachment_name:
        handler.send_header("Content-Disposition", f'attachment; filename="{attachment_name}"')
    handler.end_headers()
    handler.wfile.write(body)


class XiaomiReportHandler(BaseHTTPRequestHandler):
    server_version = "XiaomiReportEngine/1.0"

    def do_GET(self) -> None:
        path = unquote(urlparse(self.path).path)

        if path == "/":
            serve_file(self, STATIC_DIR / "index.html", "text/html; charset=utf-8")
            return

        if path == "/api/status":
            write_json(
                self,
                {
                    "app_name": APP_NAME,
                    "defaults": {
                        "service": file_status(DEFAULT_SERVICE_FILE),
                        "axio": file_status(DEFAULT_AXIO_FILE),
                        "retail": file_status(DEFAULT_RETAIL_FILE),
                        "service_master": file_status(DEFAULT_SERVICE_MASTER_FILE),
                        "channel_master": file_status(DEFAULT_CHANNEL_MASTER_FILE),
                    },
                    "outputs": {
                        "final_report": file_status(FINAL_REPORT_FILE),
                        "zonal_report": file_status(ZONAL_REPORT_FILE),
                        "channel_report": file_status(CHANNEL_REPORT_FILE),
                    },
                },
            )
            return

        if path == "/download/final_report.xlsx":
            serve_file(
                self,
                FINAL_REPORT_FILE,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "final_report.xlsx",
            )
            return

        if path == "/download/zonal_report.xlsx":
            serve_file(
                self,
                ZONAL_REPORT_FILE,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "zonal_report.xlsx",
            )
            return

        if path == "/download/final_channel_report.xlsx":
            serve_file(
                self,
                CHANNEL_REPORT_FILE,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "final_channel_report.xlsx",
            )
            return

        if path.startswith("/static/"):
            relative_path = path.removeprefix("/static/")
            static_path = (STATIC_DIR / relative_path).resolve()
            if STATIC_DIR.resolve() not in static_path.parents and static_path != STATIC_DIR.resolve():
                write_json(self, {"error": "Invalid static path."}, HTTPStatus.BAD_REQUEST)
                return
            serve_file(self, static_path)
            return

        write_json(self, {"error": "Route not found."}, HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        if path not in {"/api/generate", "/api/download"}:
            write_json(self, {"error": "Route not found."}, HTTPStatus.NOT_FOUND)
            return

        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            write_json(self, {"error": "Expected multipart form data."}, HTTPStatus.BAD_REQUEST)
            return

        try:
            if path == "/api/generate":
                result = self.handle_generate()
                write_json(self, result)
                return

            self.handle_download()
        except Exception as exc:
            write_json(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def parse_form(self) -> cgi.FieldStorage:
        return cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
            },
        )

    def resolve_report_request(
        self,
        form: cgi.FieldStorage,
        temp_path: Path,
    ) -> tuple[str, dict[str, Path]]:
        report_type = self.field_value(form, "report_type") or "service"

        if report_type == "service":
            service_path = self.materialise_upload(form, "service_file", temp_path) or DEFAULT_SERVICE_FILE
            master_path = self.materialise_upload(form, "master_file", temp_path) or DEFAULT_SERVICE_MASTER_FILE

            if not service_path.exists():
                raise FileNotFoundError(f"Daily service report not found: {service_path.name}")
            if not master_path.exists():
                raise FileNotFoundError(f"Service master not found: {master_path.name}")

            return (
                "service",
                {
                    "service_path": service_path,
                    "service_master_path": master_path,
                },
            )

        if report_type == "channel":
            axio_path = self.materialise_upload(form, "axio_file", temp_path) or DEFAULT_AXIO_FILE
            retail_path = self.materialise_upload(form, "retail_file", temp_path) or DEFAULT_RETAIL_FILE
            master_path = self.materialise_upload(form, "master_file", temp_path) or DEFAULT_CHANNEL_MASTER_FILE

            if not axio_path.exists():
                raise FileNotFoundError(f"AXIO report not found: {axio_path.name}")
            if not retail_path.exists():
                raise FileNotFoundError(f"Retail report not found: {retail_path.name}")
            if not master_path.exists():
                raise FileNotFoundError(f"Channel master not found: {master_path.name}")

            return (
                "channel",
                {
                    "axio_path": axio_path,
                    "retail_path": retail_path,
                    "channel_master_path": master_path,
                },
            )

        raise ValueError("Choose Service or Channel report.")

    def handle_generate(self) -> dict[str, object]:
        form = self.parse_form()

        with tempfile.TemporaryDirectory(prefix="xiaomi-report-") as temp_dir:
            temp_path = Path(temp_dir)
            report_type, report_kwargs = self.resolve_report_request(form, temp_path)
            return generate_reports(report_type, **report_kwargs)

    def handle_download(self) -> None:
        form = self.parse_form()
        download_key = self.field_value(form, "download_key")
        if download_key is None:
            raise ValueError("Choose a download file.")

        with tempfile.TemporaryDirectory(prefix="xiaomi-report-") as temp_dir:
            temp_path = Path(temp_dir)
            report_type, report_kwargs = self.resolve_report_request(form, temp_path)
            report_payload = generate_reports(report_type, **report_kwargs)
            available_downloads = report_payload.get("downloads", {})
            if download_key not in available_downloads:
                raise ValueError("Download is not available for the selected report.")

            download_targets = {
                "final_report": (FINAL_REPORT_FILE, "final_report.xlsx"),
                "zonal_report": (ZONAL_REPORT_FILE, "zonal_report.xlsx"),
                "channel_report": (CHANNEL_REPORT_FILE, "final_channel_report.xlsx"),
            }
            path, attachment_name = download_targets[download_key]
            serve_file(
                self,
                path,
                EXCEL_CONTENT_TYPE,
                attachment_name,
            )

    @staticmethod
    def field_value(form: cgi.FieldStorage, field_name: str) -> str | None:
        if field_name not in form:
            return None
        field = form[field_name]
        if isinstance(field, list):
            field = field[0]
        value = getattr(field, "value", None)
        if value is None:
            return None
        cleaned = str(value).strip()
        return cleaned or None

    @staticmethod
    def materialise_upload(form: cgi.FieldStorage, field_name: str, temp_path: Path) -> Path | None:
        if field_name not in form:
            return None

        field = form[field_name]
        if isinstance(field, list):
            field = field[0]

        if not getattr(field, "filename", None):
            return None

        suffix = Path(field.filename).suffix
        destination = temp_path / f"{field_name}{suffix}"
        with destination.open("wb") as output:
            field.file.seek(0)
            shutil.copyfileobj(field.file, output)
        return destination

    def log_message(self, format: str, *args: object) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {self.address_string()} - {format % args}")


def run(host: str, port: int) -> None:
    server = ThreadingHTTPServer((host, port), XiaomiReportHandler)
    print(f"{APP_NAME} running at http://{host}:{port}")
    print("Press Ctrl+C to stop.")
    server.serve_forever()


handler = XiaomiReportHandler
app = handler
application = handler


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=APP_NAME)
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=8000, type=int)
    arguments = parser.parse_args()
    try:
        run(arguments.host, arguments.port)
    except KeyboardInterrupt:
        print("\nServer stopped.")
