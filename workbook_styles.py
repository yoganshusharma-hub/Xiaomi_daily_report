from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


TITLE_FILL = PatternFill("solid", fgColor="ED6B2A")
HEADER_FILL = PatternFill("solid", fgColor="A9CBE8")
TOTAL_FILL = PatternFill("solid", fgColor="FFF3E8")
GRAND_TOTAL_FILL = PatternFill("solid", fgColor="F15A24")
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
NUMBER_FORMAT = "#,##0"


def report_title() -> str:
    return f"Xiaomi DSR 1st {datetime.now().strftime('%B')} To Till Date"


def style_service_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.title = "Daily Report"
    worksheet.insert_rows(1)

    _set_row_height(worksheet)

    worksheet["A1"] = report_title()
    _style_range(worksheet, 1, 1, 1, 5, fill=TITLE_FILL, bold=True, alignment=CENTER_ALIGN)
    worksheet.merge_cells("A1:E1")

    _style_range(worksheet, 2, 2, 1, 5, fill=HEADER_FILL, bold=True, alignment=CENTER_ALIGN)
    _style_range(worksheet, 3, worksheet.max_row, 1, 3, alignment=LEFT_ALIGN)
    _style_range(
        worksheet,
        3,
        worksheet.max_row,
        4,
        5,
        alignment=RIGHT_ALIGN,
        number_format=NUMBER_FORMAT,
    )

    for row_index in range(3, worksheet.max_row + 1):
        region = str(worksheet[f"A{row_index}"].value or "")
        state = str(worksheet[f"B{row_index}"].value or "")
        if region == "Grand Total":
            _style_range(worksheet, row_index, row_index, 1, 5, fill=GRAND_TOTAL_FILL, bold=True)
        elif region.endswith(" Total") or state.endswith(" Total"):
            _style_range(worksheet, row_index, row_index, 1, 5, fill=TOTAL_FILL, bold=True)

    _set_widths(worksheet, {"A": 18, "B": 24, "C": 42, "D": 14, "E": 16})
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:E{worksheet.max_row}"
    workbook.save(path)


def style_channel_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.insert_rows(1, amount=2)

    _set_row_height(worksheet)

    worksheet["A1"] = report_title()
    worksheet["D2"] = "AXIO"
    worksheet["F2"] = "Retail"
    worksheet["H2"] = "Total Unit"
    worksheet["I2"] = "Total GWP"

    _style_range(worksheet, 1, 1, 1, 9, fill=TITLE_FILL, bold=True, alignment=CENTER_ALIGN)
    _style_range(worksheet, 2, 3, 1, 9, fill=HEADER_FILL, bold=True, alignment=CENTER_ALIGN)
    _style_range(worksheet, 4, worksheet.max_row, 1, 3, alignment=LEFT_ALIGN)
    _style_range(
        worksheet,
        4,
        worksheet.max_row,
        4,
        9,
        alignment=RIGHT_ALIGN,
        number_format=NUMBER_FORMAT,
    )

    worksheet.merge_cells("A1:I1")
    worksheet.merge_cells("D2:E2")
    worksheet.merge_cells("F2:G2")

    for row_index in range(4, worksheet.max_row + 1):
        state = str(worksheet[f"A{row_index}"].value or "")
        distributor = str(worksheet[f"B{row_index}"].value or "")
        if state == "Grand Total":
            _style_range(worksheet, row_index, row_index, 1, 9, fill=GRAND_TOTAL_FILL, bold=True)
        elif state.endswith(" Total") or distributor.endswith(" Total"):
            _style_range(worksheet, row_index, row_index, 1, 9, fill=TOTAL_FILL, bold=True)

    _set_widths(
        worksheet,
        {"A": 22, "B": 34, "C": 52, "D": 12, "E": 14, "F": 12, "G": 14, "H": 14, "I": 14},
    )
    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = f"A3:I{worksheet.max_row}"
    workbook.save(path)


def _set_row_height(worksheet: object) -> None:
    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 23


def _set_widths(worksheet: object, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _style_range(
    worksheet: object,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
    *,
    fill: PatternFill | None = None,
    bold: bool = False,
    alignment: Alignment | None = None,
    number_format: str | None = None,
) -> None:
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
    ):
        for cell in row:
            cell.border = THIN_BORDER
            if fill is not None:
                cell.fill = fill
            if bold:
                cell.font = BOLD_FONT
            if alignment is not None:
                cell.alignment = alignment
            if number_format is not None:
                cell.number_format = number_format
